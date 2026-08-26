import base64
import json
import gzip
import http.server
import tempfile
import threading
import unittest
import urllib.parse
from datetime import date
from pathlib import Path

import phenikaa_exporter as pe
import phenikaa_login as pl


SAMPLE_EVENTS = [
    {
        "ID": "class-1",
        "TENHOCPHAN": "Học máy nâng cao",
        "TENLOPHOCPHAN": "Học máy nâng cao(N01)<br><br>",
        "NGAYHOC": "24/08/2026",
        "GIOBATDAU": 6.0,
        "PHUTBATDAU": 45.0,
        "GIOKETTHUC": 9.0,
        "PHUTKETTHUC": 25.0,
        "TIETBATDAU": 1.0,
        "TIETKETTHUC": 3.0,
        "TENPHONGHOC": "A1-601",
        "GIANGVIEN": "Phạm Tiến Lâm",
        "PHANLOAI": "LICHHOC",
    },
    {
        "ID": "exam-1",
        "TENHOCPHAN": "Kiến trúc máy tính",
        "DANGKY_LOPHOCPHAN_TEN": "Thi cuối kỳ",
        "NGAYHOC": "28/10/2026",
        "GIOBATDAU": 9.0,
        "PHUTBATDAU": 0.0,
        "GIOKETTHUC": 10.0,
        "PHUTKETTHUC": 0.0,
        "PHONGHOC_TEN": "A2-402",
        "PHANLOAI": "LICHTHI",
    },
]


class CryptoProtocolTests(unittest.TestCase):
    def test_xor_base64_round_trip_preserves_unicode(self):
        text = json.dumps({"course": "Xử lý ngôn ngữ tự nhiên"}, ensure_ascii=False)
        encoded = pe.xor_b64_encode(text, "DSA4BRINKCIpAiAPKSAv")
        self.assertEqual(pe.xor_b64_decode(encoded, "DSA4BRINKCIpAiAPKSAv"), text)

    def test_decode_bootstrap_blob_returns_required_session_fields(self):
        payload = {"userId": "student-id", "tokenJWT": "secret-token", "rootPath": "/congsinhvien"}
        blob = pe.xor_b64_encode(json.dumps(payload), "AzzS")
        self.assertEqual(pe.decode_bootstrap_blob(blob)["userId"], "student-id")
        self.assertEqual(pe.decode_bootstrap_blob(blob)["tokenJWT"], "secret-token")

    def test_parse_bootstrap_html_extracts_obfuscated_session(self):
        payload = {"userId": "student-id", "tokenJWT": "secret-token"}
        blob = pe.xor_b64_encode(json.dumps(payload), "AzzS")
        html = f'<script>AXYZCLRVN = () => "{blob}"</script>'
        self.assertEqual(pe.parse_bootstrap_html(html), payload)


class BrowserLoginCaptureTests(unittest.TestCase):
    def test_session_from_response_decodes_bootstrap_page(self):
        payload = {"userId": "student-id", "tokenJWT": "secret-token"}
        blob = pe.xor_b64_encode(json.dumps(payload), "AzzS")
        html = f'<html><script>AXYZCLRVN = () => "{blob}"</script></html>'
        self.assertEqual(pl.session_from_response("https://portal/index.aspx", html), payload)

    def test_session_from_response_ignores_pages_without_marker_or_valid_blob(self):
        self.assertIsNone(pl.session_from_response("https://portal/x", "<p>login page</p>"))
        blob = pe.xor_b64_encode('{"broken": true}', "AzzS")
        html = f'<script>AXYZCLRVN = () => "{blob}"</script>'
        self.assertIsNone(pl.session_from_response("https://portal/index.aspx", html))

    def test_token_from_authorization_strips_bearer_prefix_only(self):
        self.assertEqual(pl.token_from_authorization("Bearer abc.def.ghi"), "abc.def.ghi")
        self.assertIsNone(pl.token_from_authorization("bearer abc"))
        self.assertIsNone(pl.token_from_authorization("Bearer "))
        self.assertIsNone(pl.token_from_authorization(None))


class EventTests(unittest.TestCase):
    def test_normalize_events_sorts_and_deduplicates_by_identity(self):
        result = pe.normalize_events([SAMPLE_EVENTS[1], SAMPLE_EVENTS[0], SAMPLE_EVENTS[0]])
        self.assertEqual([e["ID"] for e in result], ["class-1", "exam-1"])
        self.assertEqual(result[0]["TENLOPHOCPHAN"], "Học máy nâng cao(N01)")

    def test_validate_date_range_rejects_reverse_range(self):
        with self.assertRaisesRegex(ValueError, "start date"):
            pe.validate_date_range(date(2026, 10, 1), date(2026, 8, 1))


class ApiAndCacheTests(unittest.TestCase):
    def test_extract_auth_from_chromium_cache_decodes_authenticated_page(self):
        payload = {"userId": "student-id", "tokenJWT": "secret-token"}
        blob = pe.xor_b64_encode(json.dumps(payload), "AzzS")
        html = f'<html><script>AXYZCLRVN = () => "{blob}"</script></html>'.encode()
        with tempfile.TemporaryDirectory() as td:
            cache_file = Path(td) / "entry_0"
            cache_file.write_bytes(b"cache-prefix\0https://portal/index.aspx\0" + gzip.compress(html) + b"cache-metadata")
            self.assertEqual(pe.extract_auth_from_cache(Path(td)), payload)

    def test_fetch_calendar_posts_encrypted_payload_and_decrypts_response(self):
        received = {}
        response_body = json.dumps({"Success": True, "Data": {"B": pe.xor_b64_encode(json.dumps(SAMPLE_EVENTS), "AzzSystem")}}).encode()

        class Handler(http.server.BaseHTTPRequestHandler):
            def do_POST(self):
                length = int(self.headers["Content-Length"])
                form = urllib.parse.parse_qs(self.rfile.read(length).decode())
                decoded = json.loads(pe.xor_b64_decode(form["A"][0], pe.CALENDAR_ACTION_KEY))
                received.update({"path": self.path, "auth": self.headers.get("Authorization"), "payload": decoded})
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(response_body)))
                self.end_headers()
                self.wfile.write(response_body)

            def log_message(self, *_args):
                pass

        server = http.server.ThreadingHTTPServer(("127.0.0.1", 0), Handler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            events = pe.fetch_calendar(
                {"userId": "student-id", "tokenJWT": "secret-token"},
                date(2026, 8, 1),
                date(2026, 10, 31),
                base_url=f"http://127.0.0.1:{server.server_port}",
            )
        finally:
            server.shutdown()
            thread.join()
            server.server_close()
        self.assertEqual(len(events), 2)
        self.assertEqual(received["auth"], "Bearer secret-token")
        self.assertEqual(received["payload"]["strNgayBatDau"], "01/08/2026")
        self.assertTrue(received["path"].endswith(pe.CALENDAR_ACTION))
    def test_cli_exports_xlsx_ics_and_source_json(self):
        response_body = json.dumps({"Success": True, "Data": {"B": pe.xor_b64_encode(json.dumps(SAMPLE_EVENTS), "AzzSystem")}}).encode()

        class Handler(http.server.BaseHTTPRequestHandler):
            def do_POST(self):
                self.rfile.read(int(self.headers["Content-Length"]))
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(response_body)))
                self.end_headers()
                self.wfile.write(response_body)

            def log_message(self, *_args):
                pass

        server = http.server.ThreadingHTTPServer(("127.0.0.1", 0), Handler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            with tempfile.TemporaryDirectory() as td:
                root = Path(td)
                auth = root / "auth.json"
                auth.write_text(json.dumps({"userId": "student-id", "tokenJWT": "secret-token"}))
                result = pe.main([
                    "--start", "2026-08-01", "--end", "2026-10-31",
                    "--auth-json", str(auth), "--out-dir", str(root / "out"),
                    "--prefix", "semester", "--base-url", f"http://127.0.0.1:{server.server_port}",
                ])
                self.assertEqual(result, 0)
                self.assertTrue((root / "out" / "semester.xlsx").exists())
                self.assertTrue((root / "out" / "semester.ics").exists())
                self.assertEqual(len(json.loads((root / "out" / "semester.json").read_text())), 2)
        finally:
            server.shutdown()
            thread.join()
            server.server_close()


class FormatTests(unittest.TestCase):
    def test_ics_contains_one_timezone_aware_event_per_record(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "calendar.ics"
            pe.write_ics(path, SAMPLE_EVENTS, calendar_name="Test")
            raw = path.read_bytes()
            text = raw.decode("utf-8")
            self.assertEqual(text.count("BEGIN:VEVENT"), 2)
            self.assertIn("DTSTART;TZID=Asia/Ho_Chi_Minh:20260824T064500", text)
            self.assertIn("SUMMARY:Exam: Kiến trúc máy tính", text)
            self.assertEqual(raw.count(b"\n"), raw.count(b"\r\n"))
            self.assertLessEqual(max(len(line) for line in raw.split(b"\r\n")), 75)

    def test_xlsx_contains_calendar_and_summary_sheets(self):
        from openpyxl import load_workbook
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "calendar.xlsx"
            pe.write_xlsx(path, SAMPLE_EVENTS)
            wb = load_workbook(path, read_only=False, data_only=False)
            self.assertEqual(wb.sheetnames, ["Calendar", "Summary"])
            self.assertEqual(wb["Calendar"].max_row, 3)
            self.assertEqual(wb["Summary"]["B3"].value, 2)
            self.assertEqual(wb["Summary"]["B5"].value, 1)


if __name__ == "__main__":
    unittest.main()
