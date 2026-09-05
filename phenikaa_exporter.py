#!/usr/bin/env python3
"""Reusable Phenikaa student-calendar API and export helpers."""

from __future__ import annotations

import argparse
import base64
import binascii
import hashlib
import json
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
import zlib
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

BOOTSTRAP_KEY = "AzzS"
RESPONSE_KEY = "AzzSystem"
TIMEZONE = "Asia/Ho_Chi_Minh"
PORTAL_BASE_URL = "https://qldtbeta.phenikaa-uni.edu.vn"
CALENDAR_ACTION = "SV_ThongTin_MH/DSA4BRINKCIpAiAPKSAv"
CALENDAR_ACTION_KEY = CALENDAR_ACTION.split("/", 1)[1]
CALENDAR_API_PATH = "/sinhvienapi3/api/" + CALENDAR_ACTION


def xor_b64_encode(text: str, key: str) -> str:
    """Match the portal's AE(): XOR JavaScript characters, then UTF-8/Base64."""
    if not key:
        raise ValueError("key must not be empty")
    encrypted = "".join(chr(ord(char) ^ ord(key[index % len(key)])) for index, char in enumerate(text))
    return base64.b64encode(encrypted.encode("utf-8")).decode("ascii")


def xor_b64_decode(encoded: str, key: str) -> str:
    """Match the portal's AD(): Base64/UTF-8 decode, then XOR characters."""
    if not key:
        raise ValueError("key must not be empty")
    encrypted = base64.b64decode(encoded).decode("utf-8")
    return "".join(chr(ord(char) ^ ord(key[index % len(key)])) for index, char in enumerate(encrypted))


def decode_bootstrap_blob(blob: str) -> dict[str, Any]:
    try:
        session = json.loads(xor_b64_decode(blob, BOOTSTRAP_KEY))
    except (binascii.Error, UnicodeError, json.JSONDecodeError) as error:
        raise ValueError("bootstrap data is invalid") from error
    if not isinstance(session, dict):
        raise ValueError("bootstrap data is invalid")
    missing = {"userId", "tokenJWT"} - session.keys()
    if missing:
        raise ValueError(f"bootstrap data is missing: {', '.join(sorted(missing))}")
    return session


def parse_bootstrap_html(html: str) -> dict[str, Any]:
    match = re.search(r'AXYZCLRVN\s*=\s*\(\)\s*=>\s*["\']([^"\']+)["\']', html)
    if not match:
        raise ValueError("AXYZCLRVN bootstrap blob was not found in the authenticated HTML")
    return decode_bootstrap_blob(match.group(1))


def extract_auth_from_cache(cache_dir: Path | str) -> dict[str, Any]:
    """Recover the newest authenticated bootstrap HTML from a Chromium disk cache.

    Chromium cache entries contain a URL/header prefix followed by a gzip response
    stream and binary metadata. This scans locally and never prints the token.
    """
    directory = Path(cache_dir).expanduser()
    if not directory.is_dir():
        raise ValueError(f"cache directory does not exist: {directory}")
    candidates = sorted((path for path in directory.iterdir() if path.is_file()), key=lambda path: path.stat().st_mtime, reverse=True)
    for path in candidates:
        try:
            data = path.read_bytes()
        except OSError:
            continue
        if b"index.aspx" not in data and b"AXYZCLRVN" not in data:
            continue
        positions = [match.start() for match in re.finditer(b"\x1f\x8b\x08", data)]
        for position in positions:
            try:
                decompressor = zlib.decompressobj(16 + zlib.MAX_WBITS)
                html = (decompressor.decompress(data[position:]) + decompressor.flush()).decode("utf-8", "replace")
                if "AXYZCLRVN" in html:
                    return parse_bootstrap_html(html)
            except (ValueError, zlib.error, UnicodeError, json.JSONDecodeError):
                continue
        if b"AXYZCLRVN" in data:
            try:
                return parse_bootstrap_html(data.decode("utf-8", "replace"))
            except ValueError:
                pass
    raise ValueError("no authenticated Phenikaa bootstrap page was found in the Chromium cache")


def fetch_calendar(
    session: dict[str, Any],
    start: date,
    end: date,
    *,
    base_url: str = PORTAL_BASE_URL,
    timeout: float = 120,
) -> list[dict[str, Any]]:
    """Fetch and decrypt calendar events for an inclusive date range."""
    validate_date_range(start, end)
    user_id = session.get("userId")
    token = session.get("tokenJWT")
    if not user_id or not token:
        raise ValueError("session requires userId and tokenJWT")
    payload = {
        "action": CALENDAR_ACTION,
        "func": "pkg_congthongtin_hssv_thongtin.LayDSLichCaNhan",
        "iM": RESPONSE_KEY,
        "strQLSV_NguoiHoc_Id": user_id,
        "strNgayBatDau": start.strftime("%d/%m/%Y"),
        "strNgayKetThuc": end.strftime("%d/%m/%Y"),
        "strChucNang_Id": "",
        "strNguoiThucHien_Id": user_id,
    }
    encrypted = xor_b64_encode(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), CALENDAR_ACTION_KEY)
    request = urllib.request.Request(
        base_url.rstrip("/") + CALENDAR_API_PATH,
        data=urllib.parse.urlencode({"A": encrypted}).encode("ascii"),
        method="POST",
        headers={
            "Authorization": "Bearer " + str(token),
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Accept": "application/json",
            "Origin": PORTAL_BASE_URL,
            "Referer": PORTAL_BASE_URL + "/congsinhvien/index.aspx",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            result = json.loads(response.read())
    except urllib.error.HTTPError as error:
        if error.code == 401:
            raise PermissionError("Phenikaa session expired; log in again and refresh the authentication source") from error
        detail = error.read(1000).decode("utf-8", "replace")
        raise RuntimeError(f"Phenikaa API returned HTTP {error.code}: {detail}") from error
    if not result.get("Success"):
        raise RuntimeError("Phenikaa API error: " + str(result.get("Message") or "unknown error"))
    data = result.get("Data")
    if isinstance(data, dict) and data.get("B"):
        data = json.loads(xor_b64_decode(data["B"], RESPONSE_KEY))
    if not isinstance(data, list):
        raise RuntimeError("Phenikaa API returned an unexpected calendar payload")
    return normalize_events(data)


def clean_html_breaks(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"<br\s*/?>", "", str(value), flags=re.IGNORECASE).strip()


def _integer(value: Any) -> int:
    return int(float(value or 0))


def event_datetime(event: dict[str, Any], *, end: bool = False) -> datetime:
    result = datetime.strptime(event["NGAYHOC"], "%d/%m/%Y")
    hour_key = "GIOKETTHUC" if end else "GIOBATDAU"
    minute_key = "PHUTKETTHUC" if end else "PHUTBATDAU"
    return result.replace(hour=_integer(event.get(hour_key)), minute=_integer(event.get(minute_key)))


def normalize_events(events: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()
    for original in events:
        event = dict(original)
        event["TENLOPHOCPHAN"] = clean_html_breaks(event.get("TENLOPHOCPHAN"))
        identity = (
            event.get("ID") or "",
            event.get("NGAYHOC"),
            event.get("GIOBATDAU"),
            event.get("PHUTBATDAU"),
            event.get("TENHOCPHAN"),
            event.get("TENLOPHOCPHAN"),
        )
        if identity not in seen:
            seen.add(identity)
            output.append(event)
    output.sort(key=lambda item: (event_datetime(item), clean_html_breaks(item.get("TENHOCPHAN"))))
    return output


def validate_date_range(start: date, end: date) -> None:
    if start > end:
        raise ValueError("start date must not be after end date")


def _ics_escape(value: Any) -> str:
    return str(value).replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\r", "").replace("\n", "\\n")


def _ics_fold(line: str) -> str:
    remaining = line.encode("utf-8")
    parts: list[str] = []
    while len(remaining) > 75:
        cut = 75
        while cut > 0 and (remaining[cut] & 0xC0) == 0x80:
            cut -= 1
        parts.append(remaining[:cut].decode("utf-8"))
        remaining = b" " + remaining[cut:]
    parts.append(remaining.decode("utf-8"))
    return "\r\n".join(parts)


def write_ics(path: Path | str, events: Iterable[dict[str, Any]], *, calendar_name: str = "Phenikaa Current Semester") -> Path:
    destination = Path(path)
    normalized = normalize_events(events)
    lines = [
        "BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//Phenikaa Calendar Exporter//EN",
        "CALSCALE:GREGORIAN", "METHOD:PUBLISH", f"X-WR-CALNAME:{_ics_escape(calendar_name)}",
        f"X-WR-TIMEZONE:{TIMEZONE}", "BEGIN:VTIMEZONE", f"TZID:{TIMEZONE}",
        f"X-LIC-LOCATION:{TIMEZONE}", "BEGIN:STANDARD", "TZOFFSETFROM:+0700",
        "TZOFFSETTO:+0700", "TZNAME:+07", "DTSTART:19700101T000000", "END:STANDARD", "END:VTIMEZONE",
    ]
    stamp = datetime.now().strftime("%Y%m%dT%H%M%SZ")
    for event in normalized:
        start, end = event_datetime(event), event_datetime(event, end=True)
        is_exam = event.get("PHANLOAI") == "LICHTHI"
        course = clean_html_breaks(event.get("TENHOCPHAN"))
        summary = ("Exam: " if is_exam else "") + course
        section = clean_html_breaks(event.get("TENLOPHOCPHAN") or event.get("DANGKY_LOPHOCPHAN_TEN"))
        room = clean_html_breaks(event.get("TENPHONGHOC") or event.get("PHONGHOC_TEN"))
        lecturer = clean_html_breaks(event.get("GIANGVIEN"))
        description = []
        if section:
            description.append("Class: " + section)
        if lecturer:
            description.append("Lecturer: " + lecturer)
        if event.get("TIETBATDAU") is not None:
            description.append(f"Periods: {_integer(event.get('TIETBATDAU'))}-{_integer(event.get('TIETKETTHUC'))}")
        attendance = clean_html_breaks(event.get("THONGTINCHUYENCAN"))
        if attendance:
            description.append("Attendance: " + attendance)
        uid = clean_html_breaks(event.get("ID")) or hashlib.sha256((summary + start.isoformat() + room).encode()).hexdigest()
        lines.extend([
            "BEGIN:VEVENT", f"UID:{uid}@phenikaa-calendar", f"DTSTAMP:{stamp}",
            f"DTSTART;TZID={TIMEZONE}:{start:%Y%m%dT%H%M%S}", f"DTEND;TZID={TIMEZONE}:{end:%Y%m%dT%H%M%S}",
            "SUMMARY:" + _ics_escape(summary), "LOCATION:" + _ics_escape(room),
            "DESCRIPTION:" + _ics_escape("\n".join(description)), "CATEGORIES:" + ("EXAM" if is_exam else "CLASS"),
            "STATUS:CONFIRMED", "TRANSP:OPAQUE", "END:VEVENT",
        ])
    lines.append("END:VCALENDAR")
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(("\r\n".join(_ics_fold(line) for line in lines) + "\r\n").encode("utf-8"))
    return destination


def write_xlsx(path: Path | str, events: Iterable[dict[str, Any]]) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo

    destination = Path(path)
    normalized = normalize_events(events)
    if not normalized:
        raise ValueError("cannot create an XLSX export with no events")
    workbook = Workbook()
    calendar = workbook.active
    calendar.title = "Calendar"
    headers = ["Date", "Weekday", "Start", "End", "Type", "Course", "Class section", "Room / Online", "Lecturer", "Periods", "Session", "Attendance", "Event ID"]
    calendar.append(headers)
    weekdays = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ nhật"]
    for event in normalized:
        start, end = event_datetime(event), event_datetime(event, end=True)
        periods = f"{_integer(event.get('TIETBATDAU'))}-{_integer(event.get('TIETKETTHUC'))}" if event.get("TIETBATDAU") is not None else ""
        calendar.append([
            start.date(), weekdays[start.weekday()], start.time(), end.time(), "Exam" if event.get("PHANLOAI") == "LICHTHI" else "Class",
            clean_html_breaks(event.get("TENHOCPHAN")), clean_html_breaks(event.get("TENLOPHOCPHAN") or event.get("DANGKY_LOPHOCPHAN_TEN")),
            clean_html_breaks(event.get("TENPHONGHOC") or event.get("PHONGHOC_TEN")), clean_html_breaks(event.get("GIANGVIEN")), periods,
            clean_html_breaks(event.get("BUOIHOC")), clean_html_breaks(event.get("THONGTINCHUYENCAN")), clean_html_breaks(event.get("ID")),
        ])
    navy, exam_fill = "17365D", "FCE4D6"
    thin = Side(style="thin", color="D9E1F2")
    for cell in calendar[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    calendar.freeze_panes = "A2"
    calendar.auto_filter.ref = calendar.dimensions
    for row in calendar.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[0].number_format, row[2].number_format, row[3].number_format = "dd/mm/yyyy", "hh:mm", "hh:mm"
        if row[4].value == "Exam":
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=exam_fill)
    for column, width in {"A": 13, "B": 13, "C": 9, "D": 9, "E": 10, "F": 36, "G": 44, "H": 24, "I": 24, "J": 10, "K": 12, "L": 18, "M": 34}.items():
        calendar.column_dimensions[column].width = width
    table = Table(displayName="SemesterCalendar", ref=f"A1:M{calendar.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    calendar.add_table(table)

    summary = workbook.create_sheet("Summary")
    summary.append(["Phenikaa calendar export", "Value"])
    summary.append(["Date range", f"{event_datetime(normalized[0]):%d/%m/%Y} – {event_datetime(normalized[-1]):%d/%m/%Y}"])
    summary.append(["Total events", len(normalized)])
    summary.append(["Classes", sum(event.get("PHANLOAI") != "LICHTHI" for event in normalized)])
    summary.append(["Exams", sum(event.get("PHANLOAI") == "LICHTHI" for event in normalized)])
    summary.append(["Courses", len({clean_html_breaks(event.get("TENHOCPHAN")) for event in normalized})])
    summary.append([])
    summary.append(["Course", "Meetings"])
    for course, count in sorted(Counter(clean_html_breaks(event.get("TENHOCPHAN")) for event in normalized).items()):
        summary.append([course, count])
    for row_number in (1, 8):
        for cell in summary[row_number]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color="FFFFFF", bold=True)
    summary.column_dimensions["A"].width, summary.column_dimensions["B"].width = 48, 24
    summary.freeze_panes = "A2"
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def _iso_date(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as error:
        raise argparse.ArgumentTypeError("dates must use YYYY-MM-DD") from error


def _load_session(args: argparse.Namespace) -> dict[str, Any]:
    if args.auth_json:
        session = json.loads(Path(args.auth_json).expanduser().read_text(encoding="utf-8"))
        missing = {"userId", "tokenJWT"} - session.keys()
        if missing:
            raise ValueError(f"authentication JSON is missing: {', '.join(sorted(missing))}")
        return session
    if args.bootstrap_html:
        return parse_bootstrap_html(Path(args.bootstrap_html).expanduser().read_text(encoding="utf-8"))
    if args.browser_login:
        import phenikaa_login

        session = phenikaa_login.login_flow(args.profile_dir)
        saved = phenikaa_login.save_auth_json(session)
        print(f"Credentials captured and saved to {saved} (mode 600).", file=sys.stderr)
        return session
    return extract_auth_from_cache(args.cache_dir)


def export_calendar_files(
    session: dict[str, Any],
    start: date,
    end: date,
    output_dir: Path | str,
    prefix: str,
    calendar_name: str,
    *,
    base_url: str = PORTAL_BASE_URL,
) -> dict[str, Any]:
    validate_date_range(start, end)
    events = fetch_calendar(session, start, end, base_url=base_url)
    if not events:
        raise RuntimeError("the API returned no calendar events for the requested range")
    destination_dir = Path(output_dir).expanduser()
    destination_dir.mkdir(parents=True, exist_ok=True)
    json_path = destination_dir / f"{prefix}.json"
    xlsx_path = destination_dir / f"{prefix}.xlsx"
    ics_path = destination_dir / f"{prefix}.ics"
    json_path.write_text(json.dumps(events, ensure_ascii=False, indent=2), encoding="utf-8")
    write_xlsx(xlsx_path, events)
    write_ics(ics_path, events, calendar_name=calendar_name)
    return {
        "events": len(events),
        "classes": sum(event.get("PHANLOAI") != "LICHTHI" for event in events),
        "exams": sum(event.get("PHANLOAI") == "LICHTHI" for event in events),
        "date_start": event_datetime(events[0]).date().isoformat(),
        "date_end": event_datetime(events[-1]).date().isoformat(),
        "json": str(json_path.resolve()),
        "xlsx": str(xlsx_path.resolve()),
        "ics": str(ics_path.resolve()),
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Export an authenticated Phenikaa learning calendar to XLSX, ICS, and JSON.")
    parser.add_argument("--start", required=True, type=_iso_date, help="Inclusive start date, YYYY-MM-DD")
    parser.add_argument("--end", required=True, type=_iso_date, help="Inclusive end date, YYYY-MM-DD")
    auth = parser.add_mutually_exclusive_group(required=True)
    auth.add_argument("--auth-json", help="Private JSON containing userId and tokenJWT")
    auth.add_argument("--bootstrap-html", help="Authenticated saved index.aspx HTML containing AXYZCLRVN")
    auth.add_argument("--cache-dir", help="Chromium Cache_Data directory containing the authenticated index page")
    auth.add_argument("--browser-login", action="store_true", help="Open a sign-in window, capture credentials into .auth.json automatically, then continue exporting")
    parser.add_argument("--out-dir", default="exports", help="Output directory (default: exports)")
    parser.add_argument("--profile-dir", default=".browser-profile", help=argparse.SUPPRESS)
    parser.add_argument("--prefix", default="phenikaa_calendar", help="Output filename prefix")
    parser.add_argument("--calendar-name", default="Phenikaa Learning Calendar", help="ICS calendar display name")
    parser.add_argument("--base-url", default=PORTAL_BASE_URL, help=argparse.SUPPRESS)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    validate_date_range(args.start, args.end)
    session = _load_session(args)
    summary = export_calendar_files(
        session,
        args.start,
        args.end,
        args.out_dir,
        args.prefix,
        args.calendar_name,
        base_url=args.base_url,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
